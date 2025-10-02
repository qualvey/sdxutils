from copy       import copy

from openpyxl                import load_workbook, Workbook
from openpyxl.styles         import Font, Border, Side,  Alignment
from openpyxl.cell.text      import InlineFont
from openpyxl.cell.rich_text import CellRichText, TextBlock, TextBlock
from openpyxl.utils import column_index_from_string, get_column_letter
from tools import  logger as mylogger
logger = mylogger.get_logger(__name__)
import shutil,os


from openpyxl                import load_workbook , Workbook
from openpyxl.cell.rich_text import  TextBlock, TextBlock

from openpyxl.styles.colors import Color

font_wenquan = Font(name='WenQuanYi Zen Hei',
                size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color=Color(rgb='FF000000')
        )

font_yahei = Font(name='Microsoft YaHei',
                size= 8,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color=Color(rgb='FF000000')
        )

yahei_inline = InlineFont(rFont='Microsoft YaHei',
                sz=11,
                color=Color(rgb='FF000000')
        )
yahei_22 = InlineFont(rFont='Microsoft YaHei',
                sz=22,
                color=Color(rgb='FFFF0000')
                    )
yahei_red_8 = InlineFont(rFont='Microsoft YaHei',
                sz=8,
                color=Color(rgb='FFFF0000')
                    )

left_alignment = Alignment(horizontal='left')
center_alignment = Alignment(horizontal='center')

class Wshandler:
    from datetime               import datetime
    def __init__(self, working_date: datetime, source_file: str, targetfile:str, data:dict ):
        
        self.source_file = source_file
        self.working_date = working_date 
        self.data = data
        self.final_data = {}
        self.targetfile = targetfile
        logger.info(f'working_date{self.working_date}')
        logger.info(f'source_file{self.source_file}')
        logger.info(f'targetfile{self.targetfile}')
         
        try:
            self.special_data = data["specialfee"]['detail']
        except (KeyError, IndexError):
            # KeyError = 根本没有 specialfee
            # IndexError = specialfee 是空 list
            self.special_data = []

    def run(self) -> Workbook:
        wb = self.init_sheet()
        self.load_data()
        self.insert_data()
        self.handle_headers()
        self.special_mark(self.ws, self.special_data, start_col='H', end_col='K')
        self.ota_comment('H')
        self.A1Bug()
        self.save()
        return wb

    def init_sheet(self) -> Workbook : 

        logger.info("init_sheet")
        working_sheetname = f'{self.working_date.month}月'
        wb = load_workbook(self.source_file)

        if working_sheetname in wb.sheetnames:
            self.ws = wb[working_sheetname]
            wb.active = self.ws
            logger.info(f'操作的worksheet{wb}:{self.ws}')
            self.ws['G37'].font = font_yahei
            self.wb = wb
            return wb
        else:
            logger.error(f'当前月份的工作表不存在，请手动检查.{working_sheetname}')
            raise ValueError("出错了")

    def special_mark(self,ws, special_data, start_col, end_col, start_row=37, end_row=47):

        thin    = Side(border_style="thin", color="888888")
        medium  = Side(border_style="medium", color="000000")
        dotted  = Side(border_style="dotted", color="000000")

        merged_ranges = ws.merged_cells.ranges

        start_col_index = column_index_from_string(start_col)
        end_col_index   = column_index_from_string(end_col)
        merge_right_col_index = end_col_index - 1  # 合并区域去掉“元”列

        merge_left_col  = start_col
        merge_right_col = get_column_letter(merge_right_col_index)
        yuan_col        = end_col  # 最右边是“元”列
        for row in range(start_row, end_row):
            for col_letter in [start_col ]:  # 例如 H 和 K
                cell = ws[f'{col_letter}{row}']
                cell.value = None

        for row in range(start_row, end_row):
            merge_range = f'{merge_left_col}{row}:{merge_right_col}{row}'
            full_range  = f'{merge_left_col}{row}:{yuan_col}{row}'

            # 如果整行已合并，拆分后再合并前几列
            if any(str(rng) == full_range for rng in merged_ranges):
                ws.unmerge_cells(full_range)
            ws.merge_cells(merge_range)

            left_cell  = ws[f'{merge_left_col}{row}']
            yuan_cell  = ws[f'{yuan_col}{row}']
            mid_cell   = ws.cell(row=row, column=merge_right_col_index)

            yuan_cell.value = None

            left_cell.border = Border(left = thin, bottom=dotted)
            mid_cell.border  = Border(left=None, right=None, bottom=dotted)
            yuan_cell.border = Border(bottom=dotted,right=medium)

            base_font = copy(ws[f'{merge_left_col}{start_row}'].font)
            left_cell.font   = base_font
            mid_cell.font    = base_font
            yuan_cell.font   = base_font

        # 设置最后一行底部边框
        bottom = end_row - 1
        ws[f'{merge_left_col}{bottom}'].border  = Border(left=thin, bottom=dotted)
        ws[f'{merge_right_col}{bottom}'].border = Border(bottom=dotted)
        ws[f'{yuan_col}{bottom}'].border        = Border(bottom=dotted, right=medium)

        # 写入数据
        titlecell = ws.cell(row=37, column=start_col_index)
        titlecell.value = "特免明细:"
        titlecell.alignment = left_alignment

        for index, (key,value) in enumerate(special_data.items()):
            row = start_row+1 + index
            if row >= end_row:
                break
            # if '：' in item:
            #     reason, value = item.split('：', 1)
            # else:
            #     reason, value = item, ''

            text_cell = ws.cell(row=row, column=start_col_index)
            text_cell.value = key+':'+str(value)
            text_cell.alignment = left_alignment

            yuan_cell = ws.cell(row=row, column=end_col_index)
            yuan_cell.value = "元"
            yuan_cell.alignment = center_alignment

    def ota_comment(self,column):
        '''
        mt_len = self.data["meituan"]["mt_count"]
        dy_len = self.data["douyin"]["dy_count"]
        mt_good_num = self.data["meituan"]["mt_good"]
        dy_good_num = self.data["douyin"]["dy_good"]
        '''
        mt_len = self.data["meituan"]["mt_count"]
        dy_len = self.data["douyin"]["dy_count"]
        mt_good_num = self.data["meituan"]["mt_good"]
        dy_good_num = self.data["douyin"]["dy_good"]
        ws = self.ws
        col_index = column_index_from_string(column)

        title  = ws.cell(row = 53, column = col_index)
        dycell = ws.cell(row = 54, column = col_index)
        mtcell = ws.cell(row = 55, column = col_index)
        good_rate_cell = ws.cell(row = 56, column = col_index)
        bad_rate_cell  = ws.cell(row = 57, column = col_index)

        for row in [54, 55, 56, 57]:
            ws.cell(row=row, column=col_index).alignment = left_alignment

        title.value  = f'OTA平台线上评价：'
        title.alignment = left_alignment

        mtcell.value = f"美团：{mt_len}单  好评数：{mt_good_num}个"
        dycell.value = f'抖音：{dy_len}单  好评数：{dy_good_num}个'
        好评率 = (mt_good_num+dy_good_num)*100/(mt_len+dy_len)
        good_rate_cell.value = f"今日好评率: {round(好评率, 2)}%"
        bad_rate_cell.value  = f"今日差评数: 0个"

    def A1Bug(self):
        '''
            A1的标题样式有bug，单独执行这个函数来处理
        '''
        red_font = Font(color="FF0000")
        rich_text = CellRichText(
            [
                "网费收入",
                TextBlock(InlineFont(red_font), "(网费+提现+找零)"),  # 使用 InlineFont 包装
            ]
        )
        self.ws['A1'] = rich_text
        font1 = Font(bold=True, size=16)
        self.ws['A1'].font = font1

    def handle_headers(self):
        mcells = [ "f1", "j1", "n1", "s1", "w1", "ac1" ]
        self.ws['A36'].number_format = 'yyyy-mm-dd'

        for cell in mcells:
            text = self.ws[cell].value
            start_index = text.find("(")
            end_index = text.find(")")

            if start_index != -1 and end_index != -1:
                # 创建 Font 对象，设置字体大小
                small_font = Font(size=8)
                # 创建 CellRichText 对象
                rich_text = CellRichText(
                    [
                        text[:start_index],  # 括号前的文本
                        TextBlock(InlineFont(small_font), text[start_index:end_index + 1]),  # 括号内的文本，小字
                        text[end_index + 1:],  # 括号后的文本
                    ]
                )
            else:
                # 如果没有括号，直接设置文本
                rich_text = CellRichText([text])
            self.ws[cell] = rich_text
            self.A1Bug()

    from openpyxl.worksheet.worksheet import Worksheet
    from datetime import datetime, date

    def find_missing_dates(self,  col_date="A", col_data="B") -> list[date]:
        from datetime import datetime
        """
        在指定 worksheet 中查找截止日期之前缺失数据的日期

        参数:
            ws (Worksheet): openpyxl 工作表对象
            end_datetime (datetime): 截止日期（只检查早于它的日期）
            col_date (str): 日期所在列 (默认 "A")
            col_data (str): 数据所在列 (默认 "B")

        返回:
            list[datetime.date]: 缺失数据的日期列表
        """
        ws = self.ws
        end_datetime = self.working_date
        
        missing_dates = []

        for row in range(2, ws.max_row + 1):  # 假设第1行是表头
            date_cell = ws[f"{col_date}{row}"].value
            data_cell = ws[f"{col_data}{row}"].value

            if date_cell is None:
                continue  # 跳过空日期

            # 转换日期
            if isinstance(date_cell, datetime):
                date_value = date_cell.date()
            else:
                # 尝试解析字符串
                try:
                    date_value = datetime.strptime(str(date_cell), "%Y-%m-%d").date()
                except:
                    continue

            # 判断是否早于 end_datetime，且数据为空
            if date_value < end_datetime.date() and (data_cell is None or str(data_cell).strip() == ""):
                missing_dates.append(date_value)

        return missing_dates

    def load_data(self)  -> dict:
        """
        :param english: 英文字段组成的原始数据
        :param elec_usage: 用电量
        :param mt: 美团营收
        :param dy: 抖音营收
        :return: 映射并处理后的数据字典
        """
        cn_en_map = {
        "网费充值"  :     "amountFee",
        "提现本金"  :     "withdrawPrincipal",
        "找零"      :     "checkoutDeposit",
        "零售"      :     "retail",
        "水吧"      :     "waterBar",
        "代购"      :     "agent",
        "退款"      :     "totalRefundFee",
        "报销"      :     None,
        "在线支付"  :     "onlineIn",
        "奖励金"    :     "awardFee",
        "卡券"      :     "cardVolumeFee",
        "特免"      :     "specialFree",
        "网费消耗"  :     "totalConsumeNetworkFee",
        "上机人次"  :     "onlineTimes",
        "上机时长"  :     "duration",
        "点单率"    :     "orderRate",
        "新会员"    :     "newMember"
    }
        elec_usage = self.data.get("elecworker",0)
        mt = self.data["meituan"].get("mt_total",0)
        dy = self.data["douyin"].get("dy_total",0)
        english = self.data.get("operation",{})
        main_data: dict[str, float | None] = {
            "用电量" : elec_usage ,
            "美团"   : mt   ,
            "抖音"   : dy,
            "口碑"   : 0.001
        }

        for cn_name, eng_name in cn_en_map.items():
            main_data[cn_name] = english.get(eng_name, 0) if eng_name else 0

        if not main_data["退款"]:
            main_data["退款"] = None
        else:
            main_data["退款"]     = -main_data["退款"] 

        if main_data["上机时长"] is not None:
            main_data["上机时长"]   = round(main_data["上机时长"] /60, 2)
        bussiness_data =main_data.copy()
        self.final_data = main_data
        return bussiness_data

    def insert_data(self) -> Worksheet:
        from typing import cast, Optional, Tuple
        from decimal import Decimal
        ws = self.ws
        working_datetime = self.working_date
        machine_sum = self.data['machine_count']
        #重新导入函数，电表类变化
        from operation import ElecDataService
    
        target_row:int | None = ElecDataService.get_row_by_date(ws, working_datetime)
        target_row  = cast(int, target_row)
        
        logger.info(f'电表数据所在行号 {target_row}')

        for col in ws.iter_cols(
            min_row=2,
            max_row=2,
            min_col=1,
            max_col=29):

            header = col[0].value  # 第二行的列标题

            if header in self.final_data and target_row:
                ws.cell(row=target_row, column=col[0].column, value=self.final_data[header]) # type: ignore

        date_str = working_datetime.strftime("%Y年%m月%d日")
        #%-m 和 %-d 中的减号用于去除月份和日期中的前导零。请注意，这种用法在某些操作系统（如 Unix/Linux）上有效，但在 Windows 上可能不被支持。

        ws["C36"].value = f"芜湖张家山店{date_str}营业状况"
        ws["G36"].value = machine_sum
        return ws

    def save(self):
        print("save workbook")
        """
        保存工作簿，并备份原始文件到 .old
        :param target_path: 目标保存完整路径
        :param source_file: 原始文件路径
        """
        source_file = self.source_file
        target_path = self.targetfile
        if not hasattr(self, "wb") or self.wb is None:
            logger.error("工作簿对象不存在，请先加载或创建工作簿")
            return

        # 确保目标目录存在
        target_dir = os.path.dirname(target_path)
        try:
            os.makedirs(target_dir, exist_ok=True)
            logger.info(f"目录 '{target_dir}' 已确保存在")
        except Exception as e:
            logger.error(f"创建目录 '{target_dir}' 时发生错误: {e}")
            return

        # 备份原文件
        backup_file = f"{source_file}.old"
        try:
            if os.path.exists(source_file):
                shutil.copyfile(source_file, backup_file)
                logger.info(f"已备份原始文件到 {backup_file}")
            else:
                logger.warning(f"源文件 {source_file} 不存在，跳过备份")
        except Exception as e:
            logger.error(f"备份文件 {source_file} 出错：{e}")
            return

        # 保存工作簿到目标路径
        try:
            self.wb.save(target_path)
            logger.debug(f"成功保存工作簿到 {target_path}")
        except FileNotFoundError:
            logger.error(f"文件 {target_path} 或目录 {target_dir} 不存在")
        except Exception as e:
            logger.error(f"保存文件 {target_path} 出错：{e}")

