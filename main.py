import shutil
import os
import argparse
from datetime                import datetime,  timedelta 

from openpyxl                import load_workbook , Workbook
from openpyxl.cell.rich_text import  TextBlock, TextBlock
from openpyxl.worksheet.worksheet import Worksheet

from meituan.main       import get_meituanSum,  get_mtgood_rates
from douyin.main        import final_out, get_dygood_rate

from operation.OperationService     import resolve_operation_data
from operation          import ThirdParty
from operation          import elecdata as electron
from specialFee         import main as specialFee
from tools              import env
from tools import  logger as mylogger
logger = mylogger.get_logger(__name__)
import xlutils.xlUtil as xlutil


from typing import cast
from concurrent.futures import ThreadPoolExecutor

machian_sum = 76

from typing import cast, Optional, Tuple
from decimal import Decimal


def init_sheet(working_datetime: datetime, source_file: str) -> Workbook :
    '''
        return workbook
    '''
    logger.info("init_sheet")
    working_sheetname = f'{working_datetime.month}月'
    wb = load_workbook(source_file)

    if working_sheetname in wb.sheetnames:
        ws = wb[working_sheetname]
        wb.active = ws
        logger.info(f'操作的worksheet{wb}:{ws}')
        ws['G37'].font = xlutil.font_yahei
        return wb
    else:
        logger.error(f'当前月份的工作表不存在，请手动检查.{working_sheetname}')
        raise ValueError("出错了")

def load_data(
        elec_usage: float, 
        mt: float, dy: float,
        english: dict,
        cn_en_map: dict
        )     -> dict:
    """
    :param english: 英文字段组成的原始数据
    :param elec_usage: 用电量
    :param mt: 美团营收
    :param dy: 抖音营收
    :return: 映射并处理后的数据字典
    """
    main_data: dict[str, float | None] = {
        "用电量" : elec_usage ,
        "美团"   : mt   ,
        "抖音"   : dy
    }

    for cn_name, eng_name in cn_en_map.items():
        main_data[cn_name] = english.get(eng_name, 0) if eng_name else 0

    if not main_data["退款"]:
        main_data["退款"] = None
    else:
        main_data["退款"]     = -main_data["退款"] 

    if main_data["上机时长"] is not None:
        main_data["上机时长"]   = round(main_data["上机时长"] /60, 2)

    main_data.setdefault("口碑",0.001)

    return main_data

def insert_data(
        ws: Worksheet, 
        data: dict,
          working_datetime:datetime
          ) -> Worksheet:

    target_row:int | None = electron.get_row_by_date(ws, working_datetime)
    target_row  = cast(int, target_row)
    
    logger.info(f'电表数据所在行号 {target_row}')

    for col in ws.iter_cols(
        min_row=2,
        max_row=2,
        min_col=1,
        max_col=29):

        header = col[0].value  # 第二行的列标题

        if header in data and target_row:
            ws.cell(row=target_row, column=col[0].column, value=data[header]) # type: ignore

    date_str = working_datetime.strftime("%Y年%m月%d日")
    #%-m 和 %-d 中的减号用于去除月份和日期中的前导零。请注意，这种用法在某些操作系统（如 Unix/Linux）上有效，但在 Windows 上可能不被支持。

    ws["C36"].value = f"芜湖张家山店{date_str}营业状况"
    ws["G36"].value = machian_sum
    return ws

def save(target_path: str,source_file:str, wb: Workbook):
    #一个函数不应该依赖于全局变量
    #dir_str
    #source_file
    try:
        os.makedirs(dir_str, exist_ok=True)
        logger.info(f"目录 '{dir_str}' 创建成功")
    except Exception as e:
        logger.error(f"创建目录时发生错误: {e}")
    try:
        shutil.move(source_file, f"{source_file}.old")
        wb.save(source_file)
        wb.save(target_path)
        logger.debug(f'成功保存到  {target_path}')
    except FileNotFoundError:
        logger.error(f"文件 {target_path} 或目录 {os.path.dirname(source_file)} 不存在")
    except Exception as e:

        logger.error(f"复制文件 {target_path} 出错：{e}")

if __name__ == "__main__":

    logger.info("主程序开始启动")

    parser = argparse.ArgumentParser(description="日报表自动化套件")
    parser.add_argument("-v", "--verbose", action="store_true", help="显示详细日志")
    parser.add_argument("-dc", "--dateconfig", action="store_true", help="使用配置文件的日期")
    parser.add_argument("-now", "--today", action="store_true", help="使用今天作为日期")
    parser.add_argument("-ne", "--no-elecfrom openpyxl import Workbook", action="store_true", help="不输入电表数据")
    parser.add_argument("-d", "--date", type=str  , help="日期 (mm-dd 格式)")

    args = parser.parse_args()

    yesterday: datetime = datetime.today() - timedelta(days=1)
    working_date_str:str = yesterday.strftime('%Y-%m-%d')
    working_datetime:datetime     = datetime.combine(yesterday, datetime.min.time())

    if args.dateconfig:
        date:datetime = env.working_datetime
        working_datetime_date  = date.date()
    if args.today:
        working_datetime_date = datetime.today()
        working_datetime      = datetime.combine(working_datetime_date, datetime.min.time())
    if args.date:
        date_str = f"{year}-{args.date}"
        try:
            working_datetime:datetime = datetime.strptime(date_str, "%Y-%m-%d")
            print("解析后的 datetime:", working_datetime)
        except ValueError:
            print("日期格式错误，应为 mm-dd，例如 09-11")

    else:
        try:
            elec_usage: float = electron.get_elecUsage(working_datetime)
            if not elec_usage:
                logger.warning('电表数据获取错误，请检查')
        except Exception as e:
            logger.error(f'{e}')

    logger.info(f'当前工作的日期是: {working_datetime}')

    dir_str     = f"{env.proj_dir}/et/{working_datetime.strftime('%m%d')}日报表"
    source_file = env.source_file
    save_path   = f"{dir_str}/2025年日报表.xlsx"

    cn_en_map = {
        "网费充值"  :     "amountFee",
        "提现本金"  :     "withdrawPrincipal",
        "找零"      :     "checkoutDeposit",
        "零售"      :     "retail",
        "水吧"      :     "waterBar",
        "代购"      :     "agent",
        "退款"      :     "totalRefundFee",
        "报销"      :     None,
        "在线支付"  :     "onlineIn",
        "奖励金"    :     "awardFee",
        "卡券"      :     "cardVolumeFee",
        "特免"      :     "specialFree",
        "网费消耗"  :     "totalConsumeNetworkFee",
        "上机人次"  :     "onlineTimes",
        "上机时长"  :     "duration",
        "点单率"    :     "orderRate",
        "新会员"    :     "newMember"
    }


    with ThreadPoolExecutor(max_workers=2) as executor:
        # submit 会立即返回 Future 对象
        meituan_data = executor.submit(get_meituanSum,working_datetime)
        meituan_good = executor.submit(get_mtgood_rates,working_datetime.strftime('%Y-%m-%d'))
        douyin_data = executor.submit(final_out,working_datetime)
        douyin_good = executor.submit(get_dygood_rate,working_datetime)

        # 调用 result() 获取函数返回值（会等待线程完成）
        mt,mt_len = meituan_data.result()
        mt_good_num = meituan_good.result()
        douyindata = douyin_data.result()
        dy_good_num = douyin_good.result()

#    mt, mt_len = get_meituanSum(working_datetime)
#    douyindata = final_out(working_datetime)
#    mt_good_num = get_mtgood_rates(working_datetime.strftime('%Y-%m-%d'))
#    dy_good_num = get_dygood_rate(working_datetime)

    if not mt:
        logger.warning('美团数据没拿到')

    douyindata:Optional[Tuple[Decimal, int]] = final_out(working_datetime)
    if not douyindata or douyindata[0] is None or douyindata[1] is None:
        print(f"缺失数据！日期: {working_datetime}")
    else:
        dy, dy_len = douyindata
        print(f"抖音数据: {dy}, 长度: {dy_len}")
    english = resolve_operation_data(working_datetime)
    if not english:
        logger.warning('运营数据没拿到')

    from typing import Callable

    ota_update: Callable[[str, datetime, int], dict] = ThirdParty.ota_update
    repeat_ids:dict = ThirdParty.check_unique(working_date_str)

    d_status = 0

    for thirdtype, ids in repeat_ids.items():
        for id in ids:
            logger.info(f'正在删除{thirdtype}:{id}')
            delete_status = ThirdParty.delete(id)
            if delete_status == 1:
                logger.warning(f'删除旧数据失败，不执行新数据更新')
                d_status = 1

    if d_status == 0:
        logger.info('删除成功')
        ota_update(ota_name='DOUYIN', date_obj=working_datetime, income=float(dy))
        ota_update(ota_name='MEITUAN', date_obj=working_datetime, income=mt)

    specialFee_list, special_sum = specialFee.get_specialFee(working_datetime)        #可以排序

    wb: Workbook = init_sheet(working_datetime,source_file=source_file)

    ws  = wb.active
    if not ws:
        raise ValueError('没有活动的worksheet')
#判断specialFeesum和specialFee是否一致I#
    missing_dates:list[date] = xlutil.find_missing_dates(ws,working_datetime)
    logger.info(f'缺失数据的日期有: {missing_dates}')

    elec_usage: float = electron.get_elecUsage(working_datetime)

    data_pure:dict = load_data(elec_usage, mt, float(dy), english, cn_en_map)
    if data_pure['特免'] != special_sum:
        logger.warning(f"特免金额不匹配，请检查!!运营数据中是{data_pure['特免']},订单列表中计算出来是{special_sum}")
    else:
        logger.info(f"特免金额匹配，pass.")

    insert_data(ws, data_pure,working_datetime)

    xlutil.special_mark(ws = ws, special_data = specialFee_list, start_col='H', end_col='K' )

    mt_good_num = get_mtgood_rates(working_datetime.strftime('%Y-%m-%d'))
    dy_good_num = get_dygood_rate(working_datetime)
    xlutil.ota_comment(ws,mt_len, dy_len,mt_good_num,dy_good_num, 'H')
    xlutil.handle_headers(ws)
    save(save_path,source_file, wb)

