

from openpyxl.styles    import Font
from openpyxl.cell.text import InlineFont 
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io

from openpyxl                import load_workbook

from tools import env
from specialFee.main import get_specialFee

source_file = env.source_file

wb = load_workbook(source_file)
ws = wb.active  # 或者指定具体的工作表，例如 wb["Sheet1"]
speciaFee = get_specialFee()
file_path = '{env.proj_dir}/test/img.jpg'


def range_to_image(worksheet, cell_range):
    """
    将指定的单元格范围转换为图像数据。

    参数：
    worksheet: openpyxl 的工作表对象。
    cell_range: 字符串，表示单元格范围，例如 "A1:D10"。

    返回：
    一个包含图像数据的字节对象。
    """
    # 创建一个新的工作簿和工作表来存储选定的范围
    temp_wb = Workbook()
    temp_ws = temp_wb.active

    # 复制指定范围的单元格到新的工作表
    for row in worksheet[cell_range]:
        for cell in row:
            new_cell = temp_ws[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell._style = cell._style

    # 保存临时工作簿到字节流
    temp_stream = io.BytesIO()
    temp_wb.save(temp_stream)
    temp_stream.seek(0)

    # 使用 openpyxl 加载临时工作簿
    temp_wb = openpyxl.load_workbook(temp_stream)
    temp_ws = temp_wb.active

    # 计算图像大小
    width = height = 0
    for row in temp_ws.iter_rows():
        height += temp_ws.row_dimensions[row[0].row].height or 15  # 默认行高为 15
        width = max(width, sum(temp_ws.column_dimensions[cell.column_letter].width or 10 for cell in row))  # 默认列宽为 10

    # 创建一个新的图像
    img = PILImage.new('RGB', (int(width * 7), int(height * 1.2)), 'white')  # 粗略估计像素尺寸
    img_draw = PILImageDraw.Draw(img)

    # 绘制单元格内容到图像
    for row in temp_ws.iter_rows():
        for cell in row:
            x = sum(temp_ws.column_dimensions[col].width or 10 for col in temp_ws.iter_cols(min_col=1, max_col=cell.col_idx-1))
            y = sum(temp_ws.row_dimensions[r].height or 15 for r in range(1, cell.row))
            img_draw.text((x * 7, y * 1.2), str(cell.value), fill='black')

    # 将图像保存到字节流
    img_byte_arr = io.BytesIO()
    img.save(img_byte_arr, format='PNG')
    img_byte_arr.seek(0)

    return img_byte_arr.getvalue()

def save_image_from_bytes(image_bytes, file_path):
    """
    将图像字节数据保存到指定文件。

    参数：
    image_bytes: 图像的字节数据。
    file_path: 要保存的目标文件路径，包括文件名和扩展名。
    """
    try:
        with open(file_path, 'wb') as file:
            file.write(image_bytes)
        print(f"图像已成功保存到 {file_path}")
    except Exception as e:
        print(f"保存图像时出错: {e}")


def special_mark(special_data):

    thin    = Side(border_style="thin", color="888888")
    medium  = Side(border_style="medium", color="000000")
    dash    = Side(border_style="dashed", color="000000")
    dotted  = Side(border_style="dotted", color="000000")

    merged_ranges = ws.merged_cells.ranges

    for row in range(37, 57):  # 37 到 57 行

        merge_range_GI = f'G{row}:I{row}'
        merge_range_GH = f'G{row}:H{row}'

        # 检查 G{row}:I{row} 是否已被合并
        if any(str(merged_range) == merge_range_GI for merged_range in merged_ranges):
            ws.unmerge_cells(merge_range_GI)
            ws.merge_cells(merge_range_GH)
        left  = ws[f'H{row}']
        right = ws[f'I{row}']
        weft  = ws[f'G{row}']
        weft.value  = None
        right.value = None
        weft.border  = Border(left = thin, bottom=dotted )
        left.border  = Border(bottom=dotted)
        right.border = Border(left=None, right=medium, bottom=dotted)
        right.font  = copy(ws['G37'].font)
        weft.font   = copy(ws['G37'].font)
 
    bottom = 57
    ws[f'G{bottom}'].border = Border(left=thin,bottom=medium)
    ws[f'H{bottom}'].border = Border(bottom=medium)
    ws[f'I{bottom}'].border = Border(bottom=medium,right=medium)

    for index, item in enumerate(special_data):
        row = 37 + index
        if row > 56:
            break  # 超过 G57，停止插入
        reason, value  = item.split('：')

        rich_string = CellRichText(
            b(yahei_inline, f'{reason}'),
            b(yahei_red_8, f'{value}'),
            b(yahei_red_8, 'sss')
            )
        cell = ws.cell(row=row, column=7)
        #cell.value = item
        cell = rich_string
        import pdb;pdb.set_trace()
        cell.alignment = left_alignment

        yuan = ws.cell(row=row, column=9)
        yuan.value = "元"
        yuan.alignment = center_alignment

#special_mark(speciaFee)
#wb.save(source_file)


img_steam = range_to_image(ws, 'A1:AF85')
save_image_from_bytes(img_steam)
