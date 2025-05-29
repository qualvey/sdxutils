
from openpyxl.styles import Font, PatternFill, Alignment, colors
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
from openpyxl import load_workbook

from tools import env

target_file = "/home/ryu/sharew/daily/2025年日报表.xlsx"

backup = "/home/ryu/sharew/daily/bak/back.ribao.xlsx"

wb = load_workbook(target_file)
ws = wb.active

def handle_headers(ws):
    mcells = [ "f1", "j1", "n1", "s1", "w1", "ac1" ]
    for cell in mcells:
        text = ws[cell].value
        start_index = text.find("(")
        end_index = text.find(")")

        if start_index != -1 and end_index != -1:
            # 创建 Font 对象，设置字体大小
            small_font = Font(size=8)
            # 创建 CellRichText 对象
            rich_text = CellRichText(
                [
                    text[:start_index],  # 括号前的文本
                    TextBlock(InlineFont(small_font), text[start_index:end_index + 1]),  # 括号内的文本，小字
                    text[end_index + 1:],  # 括号后的文本
                ]
            )
        else:
            # 如果没有括号，直接设置文本
            rich_text = CellRichText([text])
        ws[cell] = rich_text

def A1Bug():
    '''
        A1的标题样式有bug，单独执行这个函数来处理
    '''
    red_font = Font(color="FF0000")
    rich_text = CellRichText(
        [
            "网费收入",
            TextBlock(InlineFont(red_font), "(网费+提现+找零)"),  # 使用 InlineFont 包装
        ]
    )
    ws['A1'] = rich_text
    font1 = Font(bold=True, size=16)
    ws['A1'].font = font1


def delete_cells(ws,start_cell,end_cell):
    '''
    有效范围: B2,AB33
    '''
    for row in ws.iter_rows(min_row=ws[start_cell].row, max_row=ws[end_cell].row,
                            min_col=ws[start_cell].column, max_col=ws[end_cell].column):
        for cell in row:
            cell.value = None

delete_cells(ws, "B12", "AB14")

handle_headers(ws)
#单独处理A1的表头格式
A1Bug()
#wb.save(target_file)
wb.save(backup)
