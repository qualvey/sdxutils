from openpyxl import load_workbook
from datetime import datetime, date
import json
    #转换日期序列
from configer import configration

elecSheet = "/home/ryu/sharew/daily/2025年张家山店3月电表.xlsx"
reportSheet = "/home/ryu/sharew/daily/2025年日报表.xlsx"

configjson = configration()

raw_date = configjson['date']
ele_usage = configjson['elecUsage']


def str_to_excel_date(date_str):
    """
    将字符串格式的日期 ("YYYY-MM-DD") 转换为 Excel 日期序列号。
    
    参数:
    date_str (str) - 形如 "2025-03-04" 的日期字符串

    返回:
    float - Excel 日期序列号
    """
    dt = datetime.strptime(date_str, "%Y-%m-%d").date()  # 解析字符串为 date 对象
    excel_epoch = date(1899, 12, 30)  # Excel 起始日期
    return (dt - excel_epoch).days  # 计算天数差

date_sequence = str_to_excel_date(raw_date)

wbe = load_workbook(elecSheet)
wbr = load_workbook(reportSheet)
#data_only = True 获取数据而不是公式
#dataonly不能保存，保存就破坏公式了
wse = wbe.active
wsr = wbr.active
#print("A15,3-12",wsr['A14'].value)
#print("A15,3-12",wse['A15'].value)
#print("date_sequence",date_sequence)

raw_date = "2025-03-12"
search_date = datetime.strptime(raw_date, "%Y-%m-%d").date()
excel_datetime = wsr['A14'].value
#这个是datetime.datetime object
wse['A15'].value

excel_date = excel_datetime.date()
print(search_date == excel_date)


