
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 读取 Excel 文件
workbook = load_workbook('./et/20250316周报表.xlsx')
sheet = workbook.active  # 或者使用 workbook['SheetName'] 指定工作表

# 合并 B27:B29 单元格
sheet.merge_cells('B25:M27')
cell = sheet['B27']
cell.alignment = Alignment(wrap_text=True)

# 保存修改后的文件
workbook.save('20250316周报.xlsx')
